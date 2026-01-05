import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse
from django.shortcuts import redirect
from django.contrib import messages
from django.db.models import Sum, F
from django.db.models.functions import TruncMonth
from operations.models import DailyRecord, Maintenance, Transaction
from django.views.generic import TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import models


class MonthlyPerformanceView(LoginRequiredMixin, TemplateView):
    template_name = "analytics/monthly_report.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user

        daily_qs = (
            DailyRecord.objects.filter(user=user)
            .annotate(month=TruncMonth("date"))
            .values("month")
            .annotate(
                total_income=Sum("total_income"),
                total_km=Sum(F("end_km") - F("start_km")),
                days_count=models.Count("id"),
            )
            .order_by("-month")
        )

        ops_qs = (
            Transaction.objects.filter(
                record__user=user, type="COST", category__is_maintenance=False
            )
            .annotate(month=TruncMonth("record__date"))
            .values("month")
            .annotate(op_cost=Sum("amount"))
        )
        ops_dict = {item["month"]: item["op_cost"] for item in ops_qs}

        maint_qs = (
            Maintenance.objects.filter(user=user)
            .annotate(month=TruncMonth("date"))
            .values("month")
            .annotate(total_maint_cost=Sum("cost"))
        )
        maint_dict = {item["month"]: item["total_maint_cost"] for item in maint_qs}

        report_data = []
        for item in daily_qs:
            month = item["month"]

            income = item["total_income"] or 0
            km = item["total_km"] or 0

            op_cost = ops_dict.get(month, 0) or 0
            maint_cost = maint_dict.get(month, 0) or 0

            total_cost = op_cost + maint_cost
            profit = income - total_cost

            report_data.append(
                {
                    "month": month,
                    "days": item["days_count"],
                    "km": km,
                    "income": income,
                    "total_cost": total_cost,
                    "op_cost": op_cost,
                    "maint_cost": maint_cost,
                    "profit": profit,
                    "income_per_km": (income / km) if km > 0 else 0,
                    "cost_per_km": (total_cost / km) if km > 0 else 0,
                }
            )

        context["report"] = report_data
        return context


def export_reports_excel(request):
    """Gera um Excel (.xlsx) com dias em PT-BR e formatação contábil."""
    if not request.user.is_authenticated:
        return redirect("login")

    if not request.user.is_pro:
        messages.error(request, "Funcionalidade exclusiva para PRO.")
        return redirect("analytics_monthly")

    WEEKDAYS = {
        0: "Segunda-feira",
        1: "Terça-feira",
        2: "Quarta-feira",
        3: "Quinta-feira",
        4: "Sexta-feira",
        5: "Sábado",
        6: "Domingo",
    }

    records = (
        DailyRecord.objects.filter(user=request.user, is_active=False)
        .order_by("-date")
        .prefetch_related("transactions", "vehicle")
    )

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Relatório Financeiro"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="1e293b", end_color="1e293b", fill_type="solid"
    )
    center_align = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = [
        "Data",
        "Dia Semana",
        "Veículo",
        "Placa",
        "KM Inicial",
        "KM Final",
        "Rodagem (km)",
        "Receita",
        "Combustível",
        "Manutenção",
        "Outros",
        "Custo Total",
        "Lucro Líquido",
        "R$/km",
    ]
    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    row_num = 2
    for record in records:
        fuel_cost = sum(
            t.amount for t in record.transactions.all() if t.category.is_fuel
        )
        maint_cost = sum(
            t.amount for t in record.transactions.all() if t.category.is_maintenance
        )
        other_cost = record.total_cost - (fuel_cost + maint_cost)

        income_per_km = (record.profit / record.km_driven) if record.km_driven else 0

        row = [
            record.date,
            WEEKDAYS[record.date.weekday()],
            record.vehicle.model_name,
            record.vehicle.plate,
            record.start_km,
            record.end_km,
            record.km_driven,
            float(record.total_income),
            float(fuel_cost),
            float(maint_cost),
            float(other_cost),
            float(record.total_cost),
            float(record.profit),
            float(round(income_per_km, 2)),
        ]
        sheet.append(row)

        date_cell = sheet.cell(row=row_num, column=1)
        date_cell.number_format = "DD/MM/YYYY"
        date_cell.alignment = center_align

        currency_format = "R$ #,##0.00;[Red]-R$ #,##0.00"

        for col_idx in range(8, 15):
            cell = sheet.cell(row=row_num, column=col_idx)
            cell.number_format = currency_format

        profit_cell = sheet.cell(row=row_num, column=13)
        if record.profit < 0:
            profit_cell.font = Font(color="DC2626", bold=True)
        elif record.profit > 0:
            profit_cell.font = Font(color="16A34A", bold=True)

        for col in range(1, 15):
            cell = sheet.cell(row=row_num, column=col)
            cell.border = thin_border
            if col not in [3]:
                cell.alignment = center_align

        row_num += 1

    column_widths = [12, 18, 20, 12, 10, 10, 10, 15, 15, 15, 12, 15, 18, 10]
    for i, width in enumerate(column_widths, 1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        'attachment; filename="Relatorio_Financeiro_PRO.xlsx"'
    )
    workbook.save(response)

    return response
